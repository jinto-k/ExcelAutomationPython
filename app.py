import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = 0.9 * cell.value

        corrected_price_new_cell = sheet.cell(row, 4)
        corrected_price_new_cell.value = corrected_price

    sheet.cell(1, 4).value = 'corrected price'

    # The following code is to create a bar graph
    values = Reference(sheet,
                       min_row=2,
                       max_row=4,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)



